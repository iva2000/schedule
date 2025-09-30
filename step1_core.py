# -*- coding: utf-8 -*-
"""
step1_core.py — «ядро» Шага 1 (без I/O Colab).
Функции получают на вход байты Excel и возвращают:
 - обработанные DataFrame'ы по листам,
 - XLSX (в байтах) «…, файл обработан.xlsx» с автошириной и зеброй недель (1-я неделя серая, далее через одну).
Логика перенесена из колаб-скрипта «Шаг 1» (определение заголовка, нормализация, «зебра» и пр.).
"""

from __future__ import annotations
from typing import Dict, Tuple, List, Optional
from io import BytesIO
from collections import OrderedDict
from datetime import datetime, timedelta, date
import re
import numpy as np
import pandas as pd

# ========= Хелперы из вашего «Шаг 1» =========

def norm_txt(s):
    if s is None or (isinstance(s, float) and np.isnan(s)):
        return ""
    s = str(s).replace("ё", "е").replace("Ё", "Е").strip().lower()
    return s

def normalize_header_value(s: str) -> str:
    s = norm_txt(s)
    s = re.sub(r"[^\w]+", "_", s, flags=re.U).strip("_")
    return s or "col"

def make_unique(names):
    seen, out = {}, []
    for n in names:
        base = n
        if base not in seen:
            seen[base] = 1; out.append(base)
        else:
            seen[base] += 1; out.append(f"{base}_{seen[base]}")
    return out

KEYWORDS = {
    "дата","день","время","начало","конец","предмет","дисципл","занят",
    "преподав","тип","форма","лекц","семин","практик","аудит","каб",
    "комната","групп","поток","недел","пара","номер","формат",
    "дисциплина","специальность","группы","ауд","время","даты","тип"
}

DOW_FULL = ["понедельник","вторник","среда","четверг","пятница","суббота","воскресенье"]
DOW_SHORT = ["пн","вт","ср","чт","пт","сб","вс"]
DOW_CANON = {"пн":"понедельник","вт":"вторник","ср":"среда","чт":"четверг","пт":"пятница","сб":"суббота","вс":"воскресенье"}

DOW_SEP_RE = re.compile(
    r"^\s*(?:\d{1,2}\s*[.)-]\s*)?(понедельник|вторник|среда|четверг|пятница|суббота|воскресенье)\s*$",
    flags=re.IGNORECASE | re.UNICODE
)

def score_header_row(row_vals) -> int:
    score = 0
    for v in row_vals:
        t = norm_txt(v)
        if t:
            score += 1
            if any(kw in t for kw in KEYWORDS):
                score += 2
    return score

def detect_header_index(df: pd.DataFrame, max_scan: int = 12) -> int:
    limit = min(max_scan, len(df))
    best_idx, best_score = 0, -1
    for i in range(limit):
        s = score_header_row(list(df.iloc[i, :].values))
        if s > best_score:
            best_idx, best_score = i, s
    return best_idx

def find_first_col(df: pd.DataFrame, substrings):
    for c in df.columns:
        if any(sub in norm_txt(c).replace(" ", "_") for sub in substrings):
            return c
    return None

def split_time_like(s):
    if s is None or (isinstance(s, float) and np.isnan(s)):
        return (None, None)
    s_norm = str(s).replace("—","-").replace("–","-").replace("−","-")
    s_norm = re.sub(r"\s+","", s_norm)
    if "-" not in s_norm:
        return (None, None)
    a, b = s_norm.split("-", 1)
    def to_hhmm(p):
        m = re.match(r"^(\d{1,2})([:\.]?)(\d{0,2})$", p)
        if m:
            hh = int(m.group(1)); mm = int(m.group(3) or 0)
            if 0 <= hh < 24 and 0 <= mm < 60:
                return f"{hh:02d}:{mm:02d}"
            return None
        m2 = re.match(r"^(\d{2})(\d{2})$", p)
        if m2:
            hh = int(m2.group(1)); mm = int(m2.group(2))
            if 0 <= hh < 24 and 0 <= mm < 60:
                return f"{hh:02d}:{mm:02d}"
        return None
    return (to_hhmm(a), to_hhmm(b))

def map_day_name(x):
    t = norm_txt(x)
    if not t: return None
    if t in DOW_FULL: return t
    if t in DOW_SHORT: return DOW_CANON[t]
    for full in DOW_FULL:
        if full in t:
            return full
    return None

def ensure_columns(df: pd.DataFrame, needed: list) -> pd.DataFrame:
    for name in needed:
        if name not in df.columns:
            df[name] = None
    return df

def unique_join(values, sep=", "):
    out = OrderedDict()
    def add_one(x):
        if x is None or (isinstance(x, float) and np.isnan(x)): return
        s = str(x).strip()
        if s: out.setdefault(s, True)
    def split_many(x):
        if x is None or (isinstance(x, float) and np.isnan(x)): return []
        return [p.strip() for p in re.split(r"[,;\/\n]+", str(x)) if p.strip()]
    if isinstance(values, (list, tuple, set)):
        for v in values:
            for p in split_many(v): add_one(p)
    else:
        for p in split_many(values): add_one(p)
    return sep.join(out.keys()) if out else None

def build_groups_column(df: pd.DataFrame) -> pd.DataFrame:
    cands = [c for c in df.columns if ("групп" in norm_txt(c)) or ("поток" in norm_txt(c))]
    if "группы" not in df.columns:
        df["группы"] = None
    if not cands:
        return df
    df["группы"] = df.apply(lambda r: (unique_join([r[c] for c in cands if c in r.index]) or r.get("группы")), axis=1)
    return df

def autosize_openpyxl_sheet(ws, df: pd.DataFrame, max_width=60, padding=2):
    from openpyxl.utils import get_column_letter
    for j, col in enumerate(df.columns, start=1):
        header_len = len(str(col)) if col is not None else 0
        max_len = header_len
        for v in df[col].values:
            l = 0 if (v is None or (isinstance(v, float) and np.isnan(v))) else len(str(v))
            if l > max_len:
                max_len = l
        ws.column_dimensions[get_column_letter(j)].width = min(max_len + padding, max_width)

def unique_excel_sheet_name(name, used):
    base = (str(name) or "Sheet")[:31]
    if base not in used:
        used.add(base); return base
    i = 2
    while True:
        cand = (base[:28] + f"_{i}")[:31]
        if cand not in used:
            used.add(cand); return cand
        i += 1

def process_day_separators(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if "день_недели" not in df.columns:
        df["день_недели"] = None
    sep_rows_idx = []
    current_day = None
    for idx in df.index:
        row = df.loc[idx]
        found_day = None
        for v in row.values:
            t = norm_txt(v)
            if not t: continue
            m = DOW_SEP_RE.match(t)
            if m:
                found_day = norm_txt(m.group(1)); break
        if found_day:
            current_day = found_day; sep_rows_idx.append(idx)
        else:
            if (df.at[idx, "день_недели"] is None) or (isinstance(df.at[idx, "день_недели"], float) and np.isnan(df.at[idx, "день_недели"])):
                if current_day:
                    df.at[idx, "день_недели"] = current_day
    if sep_rows_idx:
        df = df.drop(index=sep_rows_idx).reset_index(drop=True)
    return df

def fill_day_of_week(df: pd.DataFrame) -> pd.DataFrame:
    target_col = "день_недели" if "день_недели" in df.columns else ("день" if "день" in df.columns else None)
    if target_col is None:
        df["день_недели"] = None; target_col = "день_недели"
    src_col = find_first_col(df, ["день_недел", "день"])
    if src_col:
        mapped = df[src_col].map(map_day_name)
        df[target_col] = mapped.where(mapped.notna(), df.get(target_col))
    else:
        days, cur = [], None
        for _, row in df.iterrows():
            found = None
            for v in row.values:
                m = map_day_name(v)
                if m: found = m; break
            if found: cur = found
            days.append(cur)
        df[target_col] = pd.Series(days).where(df[target_col].isna(), df[target_col])
    return df

REPLACERS = [
    (re.compile(r"удаленн*ый\s*доступ\s*\(для\s*дистанционного\s*тестирования\)\s*дистанционное\s*тестирование"), "Удал.тест."),
    (re.compile(r"науки\s*о\s*жизни"), "НоЖ"),
    (re.compile(r"физика\s*,\s*математика"), "Физ.,мат."),
    (re.compile(r"\bфизика\b"), "Физ."),
    (re.compile(r"\bинформатика\b"), "Инф."),
]

def apply_text_rules_and_ffill(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    text_cols = [c for c in df.columns if df[c].dtype == object]
    if text_cols:
        def _replace_val(x):
            if x is None or (isinstance(x, float) and np.isnan(x)): return x
            t = norm_txt(x)
            for pat, repl in REPLACERS:
                if pat.search(t): return repl
            return x
        for c in text_cols:
            df[c] = df[c].apply(_replace_val)
        df[text_cols] = df[text_cols].replace(r"^\s*$", np.nan, regex=True).ffill()
    spec_cols = []
    for c in df.columns:
        nc = norm_txt(c).replace(" ", "_")
        if nc in {"дисциплина","специальность","группы","ауд","аудитория","аудитория_комната"} or ("аудит" in nc):
            spec_cols.append(c)
    if spec_cols:
        df[spec_cols] = df[spec_cols].replace(r"^\s*$", np.nan, regex=True).ffill()
    return df

TIME_RE = re.compile(r'(?<!\d)([01]?\d|2[0-3])[:\.]([0-5]\d)')

def pick_time_source_column(df: pd.DataFrame):
    cols = list(df.columns)
    if "время" in cols: return "время"
    if "time_range" in cols: return "time_range"
    for c in cols:
        if "время" in norm_txt(c).replace(" ", "_"):
            return c
    return None

def parse_two_times_from_cell(val):
    if val is None or (isinstance(val, float) and np.isnan(val)): return (None, None)
    s = str(val).replace("—","-").replace("–","-").replace("−","-").replace(".", ":")
    matches = TIME_RE.findall(s)
    if len(matches) >= 2:
        h1, m1 = matches[0]; h2, m2 = matches[1]
        try:
            hh1, mm1 = int(h1), int(m1); hh2, mm2 = int(h2), int(m2)
            if 0 <= hh1 < 24 and 0 <= mm1 < 60 and 0 <= hh2 < 24 and 0 <= mm2 < 60:
                return (f"{hh1:02d}:{mm1:02d}", f"{hh2:02d}:{mm2:02d}")
        except Exception:
            pass
    return (None, None)

def final_extract_times(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for need_col in ["начало", "конец"]:
        if need_col not in df.columns: df[need_col] = None
    src = pick_time_source_column(df)
    if not src: return df
    starts, ends = [], []
    for v in df[src].values:
        s, e = parse_two_times_from_cell(v); starts.append(s); ends.append(e)
    s_ser = pd.Series(starts, index=df.index); e_ser = pd.Series(ends, index=df.index)
    mask_both = s_ser.notna() & e_ser.notna()
    df.loc[mask_both & df["начало"].isna(), "начало"] = s_ser[mask_both]
    df.loc[mask_both & df["конец"].isna(),  "конец"]  = e_ser[mask_both]
    return df

DATE_TOKEN_RE = re.compile(r'\b(\d{1,2})[\.\/](\d{1,2})(?:[\.\/](\d{2,4}))?\b')
RANGE_NORMALIZER_RE = re.compile(r'\b(по|до)\b', flags=re.IGNORECASE)
ITEM_SPLIT_RE = re.compile(r'[;,\/\n]+')
WEEKDAY_INDEX = {
    "понедельник":0,"вторник":1,"среда":2,"четверг":3,"пятница":4,"суббота":5,"воскресенье":6,
    "пн":0,"вт":1,"ср":2,"чт":3,"пт":4,"сб":5,"вс":6,
}
def _fmt_date(d: datetime) -> str:
    return d.strftime("%d.%m.%Y")
def _safe_parse_date(day:int, month:int, year:int):
    if year < 100: year = 2000 + year
    try: return datetime(year, month, day)
    except Exception: return None
def parse_date_token(tok: str, ref_year: int = None):
    m = DATE_TOKEN_RE.search(tok)
    if not m: return None
    d = int(m.group(1)); mth = int(m.group(2)); y = m.group(3)
    year = int(y) if y else (ref_year if ref_year else datetime.today().year)
    return _safe_parse_date(d, mth, year)
def expand_range_to_weekday_list(d1: datetime, d2: datetime, target_wd: int):
    if d1 is None or d2 is None: return []
    if d2 < d1: d1, d2 = d2, d1
    out, seen = [], set()
    def _add(x):
        s = _fmt_date(x)
        if s not in seen: out.append(s); seen.add(s)
    _add(d1); cur = d1 + timedelta(days=1)
    while cur < d2:
        if target_wd is not None and cur.weekday() == target_wd: _add(cur)
        cur += timedelta(days=1)
    _add(d2)
    return out
def expand_dates_cell(cell_val, dow_value):
    if cell_val is None or (isinstance(cell_val, float) and np.isnan(cell_val)): return cell_val
    txt = str(cell_val).replace("—","-").replace("–","-").replace("−","-").replace("…","-").replace("..","-")
    txt = RANGE_NORMALIZER_RE.sub("-", txt)
    parts = [p.strip() for p in ITEM_SPLIT_RE.split(txt) if p.strip()]
    if not parts: return cell_val
    wd = WEEKDAY_INDEX.get(norm_txt(dow_value), None)
    result_pieces = []
    for part in parts:
        tokens = DATE_TOKEN_RE.findall(part)
        if len(tokens) >= 2:
            t1 = ".".join([tokens[0][0], tokens[0][1], tokens[0][2]]) if tokens[0][2] else f"{tokens[0][0]}.{tokens[0][1]}"
            ref_year = int(tokens[0][2]) if tokens[0][2] else None
            t2 = ".".join([tokens[1][0], tokens[1][1], tokens[1][2]]) if tokens[1][2] else f"{tokens[1][0]}.{tokens[1][1]}"
            d_start = parse_date_token(t1)
            d_end   = parse_date_token(t2, ref_year=ref_year if ref_year is not None else (d_start.year if d_start else None))
            expanded = expand_range_to_weekday_list(d_start, d_end, wd)
            result_pieces.extend(expanded if expanded else [part])
        elif len(tokens) == 1:
            t = ".".join([tokens[0][0], tokens[0][1], tokens[0][2]]) if tokens[0][2] else f"{tokens[0][0]}.{tokens[0][1]}"
            year_hint = int(tokens[0][2]) if tokens[0][2] else None
            d_single = parse_date_token(t, ref_year=year_hint)
            result_pieces.append(_fmt_date(d_single) if d_single else part)
        else:
            result_pieces.append(part)
    return "; ".join(result_pieces) if result_pieces else cell_val
def expand_dates_column(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    col = "даты" if "даты" in df.columns else ("дата" if "дата" in df.columns else None)
    if not col:
        for c in df.columns:
            if "дат" in norm_txt(c): col = c; break
    if not col: return df
    day_col = "день_недели" if "день_недели" in df.columns else ("день" if "день" in df.columns else None)
    if day_col is None:
        df[col] = df[col].apply(lambda x: expand_dates_cell(x, None))
    else:
        df[col] = df.apply(lambda r: expand_dates_cell(r.get(col), r.get(day_col)), axis=1)
    return df

DATE_ONLY_TOKEN = re.compile(r'\b(\d{1,2})[\.\/](\d{1,2})(?:[\.\/](\d{2,4}))?\b')
def resolve_date_column(df: pd.DataFrame):
    if "даты" in df.columns: return "даты"
    if "дата" in df.columns: return "дата"
    for c in df.columns:
        if "дат" in norm_txt(c): return c
    return None
def parse_date_for_sort(x):
    if isinstance(x, (datetime, pd.Timestamp)): return x.date()
    if isinstance(x, date): return x
    if x is None or (isinstance(x, float) and np.isnan(x)): return None
    s = str(x); m = DATE_ONLY_TOKEN.search(s)
    if not m: return None
    d_, mth, y = int(m.group(1)), int(m.group(2)), m.group(3)
    year = int(y) if y else datetime.today().year
    if year < 100: year = 2000 + year
    try: return date(year, mth, d_)
    except Exception: return None
def parse_time_minutes(x):
    if x is None or (isinstance(x, float) and np.isnan(x)): return 24*60+59
    s = str(x); m = TIME_RE.search(s.replace(".", ":"))
    if not m: return 24*60+59
    hh, mm = int(m.group(1)), int(m.group(2))
    return hh*60 + mm if (0 <= hh < 24 and 0 <= mm < 60) else 24*60+59
def explode_dates_and_sort(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    date_col = resolve_date_column(df)
    if not date_col: return df
    rows = []
    for _, row in df.iterrows():
        val = row.get(date_col)
        if val is None or (isinstance(val, float) and np.isnan(val)) or isinstance(val, (datetime, pd.Timestamp, date)):
            rows.append(row); continue
        parts = [p.strip() for p in re.split(r"\s*;\s*", str(val)) if p.strip()]
        if len(parts) <= 1:
            rows.append(row)
        else:
            for p in parts:
                new_row = row.copy(); new_row[date_col] = p; rows.append(new_row)
    out = pd.DataFrame(rows, columns=df.columns)
    out["_sort_date"] = out[date_col].apply(parse_date_for_sort)
    out["_sort_time"] = out["начало"].apply(parse_time_minutes) if "начало" in out.columns else (24*60+59)
    out = out.sort_values(by=["_sort_date", "_sort_time"], ascending=[True, True], kind="mergesort")
    out = out.drop(columns=["_sort_date", "_sort_time"], errors="ignore").reset_index(drop=True)
    return out

DAYS_SUFFIX_RE = re.compile(r"\(\s*\d+\s*дн\.\s*\)", flags=re.IGNORECASE)
def cleanup_days_suffixes(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for c in df.columns:
        if df[c].dtype == object:
            df[c] = df[c].apply(lambda x: re.sub(DAYS_SUFFIX_RE, "", x).strip() if isinstance(x, str) else x)
    return df
def finalize_layout(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if "даты" in df.columns and "дата" not in df.columns:
        df = df.rename(columns={"даты": "дата"})
    elif "даты" in df.columns and "дата" in df.columns:
        df["дата"] = df["дата"].where(df["дата"].notna(), df["даты"]); df = df.drop(columns=["даты"])
    if "день_недели" in df.columns and "день" not in df.columns:
        df = df.rename(columns={"день_недели": "день"})
    elif "день_недели" in df.columns and "день" in df.columns:
        df["день"] = df["день"].where(df["день"].notna(), df["день_недели"]); df = df.drop(columns=["день_недели"])
    df = df.drop(columns=["время"], errors="ignore")
    first = [c for c in ["дата", "день", "начало", "конец"] if c in df.columns]
    rest = [c for c in df.columns if c not in first]
    tail = []
    if "специальность" in rest:
        rest.remove("специальность"); tail.append("специальность")
    if "дисциплина" in rest:
        rest.remove("дисциплина"); tail.append("дисциплина")
    df = df[first + rest + tail]
    df = cleanup_days_suffixes(df)
    return df

def first_non_empty(series):
    for x in series:
        if pd.notna(x) and str(x).strip() != "":
            return x
    return np.nan
def merge_groups(series):
    return unique_join(list(series), sep="; ")
def merge_same_date_start(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if "дата" not in df.columns or "начало" not in df.columns:
        return df
    if "группы" not in df.columns:
        df["группы"] = None
    original_order = list(df.columns)
    mask = df["дата"].notna() & df["начало"].notna()
    keep = df[~mask].copy()
    if mask.any():
        groupers = ["дата", "начало"]
        agg_cols = [c for c in df.columns if c not in groupers]
        agg_dict = {c: (merge_groups if c == "группы" else first_non_empty) for c in agg_cols}
        collapsed = df[mask].groupby(groupers, dropna=False, sort=False).agg(agg_dict).reset_index()
        out = pd.concat([collapsed, keep], ignore_index=True, sort=False)
    else:
        out = df
    out["_sort_date"] = out["дата"].apply(parse_date_for_sort)
    out["_sort_time"] = out["начало"].apply(parse_time_minutes) if "начало" in out.columns else (24*60+59)
    out = out.sort_values(by=["_sort_date","_sort_time"], ascending=[True, True], kind="mergesort")
    out = out.drop(columns=["_sort_date","_sort_time"], errors="ignore").reset_index(drop=True)
    ordered = [c for c in original_order if c in out.columns] + [c for c in out.columns if c not in original_order]
    out = out[ordered]
    return out

# ========= Основные API-функции =========

def process_workbook_step1(file_bytes: bytes) -> Tuple[Dict[str, pd.DataFrame], bytes]:
    """
    Вход: байты Excel (xls/xlsx).
    Возвращает:
      - processed: словарь {имя_листа: DataFrame} после всех трансформаций,
      - xlsx_bytes: XLSX «…, файл обработан.xlsx» (в байтах) c автошириной и «зеброй» недель.
    """
    # 1) Чтение книги с толерантными fallback
    buf = BytesIO(file_bytes)
    book = None; last_err = None
    for eng in ("openpyxl", None, "xlrd"):
        try:
            buf.seek(0)
            book = pd.read_excel(buf, sheet_name=None, header=None, engine=eng)
            break
        except Exception as e:
            last_err = e
    if book is None:
        raise RuntimeError(f"Не удалось прочитать книгу: {last_err}")

    # 2) Обработка листов (по «Шаг 1»)
    processed = {}
    for sheet_name, raw_df in book.items():
        try:
            df0 = raw_df.dropna(how="all").dropna(how="all", axis=1)
            if df0.empty:
                continue
            hdr_idx = detect_header_index(df0, max_scan=12)
            headers_norm = make_unique([normalize_header_value(x) for x in list(df0.iloc[hdr_idx].values)])
            df_body = df0.iloc[hdr_idx+1:].copy().reset_index(drop=True)
            n_cols = len(headers_norm)
            if df_body.shape[1] < n_cols:
                for _ in range(n_cols - df_body.shape[1]):
                    df_body[df_body.shape[1]] = None
            elif df_body.shape[1] > n_cols:
                df_body = df_body.iloc[:, :n_cols]
            df_body.columns = headers_norm
            df_body = df_body.dropna(how="all", axis=1)
            df_body = ensure_columns(df_body, ["день_недели", "начало", "конец"])
            df_body = process_day_separators(df_body)
            time_col = find_first_col(df_body, ["время"])
            if time_col:
                if len(df_body):
                    starts, ends = zip(*[split_time_like(v) for v in df_body[time_col].values])
                    df_body["начало"] = df_body["начало"].where(df_body["начало"].notna(), pd.Series(starts))
                    df_body["конец"]  = df_body["конец"].where(df_body["конец"].notna(),  pd.Series(ends))
            df_body = fill_day_of_week(df_body)
            df_body = build_groups_column(df_body)
            processed[sheet_name] = df_body
        except Exception:
            continue

    if not processed:
        raise RuntimeError("Не удалось обработать ни один лист — проверьте исходные данные.")

    # === Финальные преобразования ===
    for sname in list(processed.keys()):
        processed[sname] = apply_text_rules_and_ffill(processed[sname])
    for sname in list(processed.keys()):
        processed[sname] = final_extract_times(processed[sname])
    for sname in list(processed.keys()):
        processed[sname] = expand_dates_column(processed[sname])
    for sname in list(processed.keys()):
        processed[sname] = explode_dates_and_sort(processed[sname])
    for sname in list(processed.keys()):
        processed[sname] = finalize_layout(processed[sname])
    for sname in list(processed.keys()):
        processed[sname] = merge_same_date_start(processed[sname])

    # 3) Экспорт в XLSX (в память) с автошириной и «зеброй недель»
    from openpyxl.styles import PatternFill

    out_buf = BytesIO()
    dfs_in_order = []
    sheet_names_in_order = []

    with pd.ExcelWriter(out_buf, engine="openpyxl") as writer:
        used = set()
        for sname, df in processed.items():
            safe = unique_excel_sheet_name(str(sname) if sname is not None else "Sheet", used)
            sheet_names_in_order.append(safe)
            dfs_in_order.append(df)
            df.to_excel(writer, sheet_name=safe, index=False)
            ws = writer.sheets[safe]
            autosize_openpyxl_sheet(ws, df)

        # НОВАЯ ЗАЛИВКА: первая неделя серым, далее через одну
        fill = PatternFill(fill_type="solid", fgColor="FFEFEFEF")
        wb = writer.book

        def _parse_date_for_sort(x):
            return parse_date_for_sort(x)

        for ws_name, df in zip(sheet_names_in_order, dfs_in_order):
            ws = wb[ws_name]
            if "дата" not in df.columns:
                continue
            week_order = {}  # (year, week) -> 1..N по мере появления
            counter = 0
            for i, val in enumerate(df["дата"].tolist(), start=2):  # +1 строка заголовка
                dt = _parse_date_for_sort(val)
                if not dt:
                    continue
                iso = dt.isocalendar()
                key = (iso[0], iso[1])  # (year, week)
                if key not in week_order:
                    counter += 1
                    week_order[key] = counter
                # Нечетные недели — заливаем
                if week_order[key] % 2 == 1:
                    for cell in ws[i]:
                        cell.fill = fill

    return processed, out_buf.getvalue()


def preview_tables_html(processed: Dict[str, pd.DataFrame], n: int = 20) -> str:
    """Компактный HTML-предпросмотр head(n) по листам для Streamlit."""
    parts = [
        '<style>table{font-size:12px;border-collapse:collapse} th,td{border:1px solid #ddd;padding:4px} th{background:#f7f7f7}</style>'
    ]
    for sname, df in processed.items():
        parts.append(f"<h4>{sname}</h4>")
        parts.append(df.head(n).to_html(index=False, border=0))
    return "\n".join(parts)
