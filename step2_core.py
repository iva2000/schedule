# -*- coding: utf-8 -*-
"""
step2_core.py — «ядро» Шага 2 (без I/O Colab).
Получает XLSX (байты) из Шага 1; выполняет дополнительную обработку и создаёт:
 - XLSX (в байтах) «…, группы помечены цветом.xlsx»;
 - PNG-диаграммы (в байтах) по каждому листу.
Логика перенесена и адаптирована из вашего «Шаг 2» (палитра, «зебра», построение диаграмм). 
"""

from __future__ import annotations
from typing import Dict, Tuple, List, Optional
from io import BytesIO
import math, re, colorsys, random
from datetime import datetime, date, time, timedelta

import numpy as np
import pandas as pd

import matplotlib
matplotlib.rcParams["font.family"] = "DejaVu Sans"
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.patheffects as path_effects
from matplotlib.ticker import FixedLocator, FixedFormatter

from dateutil.relativedelta import relativedelta
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

random.seed(0)

# ===== Константы, как в «Шаг 2» =====
FIG_WIDTH_INCH  = 21
FIG_DPI         = 400
BAR_WIDTH_DAYS  = 0.8

RU_WEEKDAYS = ["Понедельник","Вторник","Среда","Четверг","Пятница","Суббота","Воскресенье"]

# ===== Хелперы «Шага 2» =====
def sanitize_fs_name(s: str) -> str:
    return re.sub(r'[\\/*?:"<>|]', '_', str(s)).strip()

def find_column(cols, variants):
    norm = lambda x: re.sub(r"\s+", " ", re.sub(r"[^\wА-Яа-яЁё]+", " ", str(x))).strip().lower()
    norm_cols = {norm(c): c for c in cols}
    for v in variants:
        key = norm(v)
        if key in norm_cols: return norm_cols[key]
    for v in variants:
        key = norm(v)
        for nc, orig in norm_cols.items():
            if key in nc: return orig
    return None

def to_date_series(s):
    out = pd.to_datetime(s, errors="coerce", dayfirst=True)
    mask_num = s.apply(lambda x: isinstance(x, (int, float))) & out.isna()
    if mask_num.any():
        out.loc[mask_num] = pd.to_datetime(s[mask_num], origin="1899-12-30", unit="D", errors="coerce")
    return out

def parse_time_cell(x):
    if pd.isna(x): return None
    if isinstance(x, pd.Timestamp): return time(x.hour, x.minute, x.second)
    if isinstance(x, time): return x
    if isinstance(x, (int, float)) and not isinstance(x, bool):
        if np.isfinite(x):
            mins = int(round(float(x) * 24 * 60))
            mins = max(0, min(24*60-1, mins))
            return (datetime(2000,1,1) + timedelta(minutes=mins)).time()
        return None
    s = str(x).strip().replace(" ", "").replace(".", ":")
    m = re.match(r"^(\d{1,2}):(\d{2})(?::(\d{2}))?$", s)
    if m:
        hh = int(m.group(1)); mm = int(m.group(2)); ss = int(m.group(3) or 0)
        hh = max(0, min(23, hh)); mm = max(0, min(59, mm)); ss = max(0, min(59, ss))
        return time(hh, mm, ss)
    try:
        ts = pd.to_datetime(s, errors="raise")
        return time(ts.hour, ts.minute, ts.second)
    except Exception:
        return None

def time_to_minutes(t: time) -> int: return t.hour*60 + t.minute + t.second//60
def minutes_to_label_custom(m: int) -> str:
    h = int(m//60)
    return "8:00" if h == 8 else f"{h}"

def month_name_ru(dt: datetime) -> str:
    names = ["Январь","Февраль","Март","Апрель","Май","Июнь","Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"]
    return names[dt.month-1]

def week_monday(d: date) -> date: return d - timedelta(days=(d.weekday()))
def daterange(d0: date, d1: date):
    cur = d0
    while cur <= d1:
        yield cur
        cur += timedelta(days=1)

def sym_offset(index: int, step: float) -> float:
    if index == 0: return 0.0
    sign = 1 if (index % 2 == 1) else -1
    magn = (index + 1)//2
    return sign * magn * step

def autosize_all_sheets(writer, sheet_to_df):
    for name, df in sheet_to_df.items():
        ws = writer.sheets[name]
        for col_idx, col in enumerate(df.columns, start=1):
            maxlen = len(str(col)) if col is not None else 0
            for v in df[col].astype(str).values:
                maxlen = max(maxlen, len(v))
            width = min(100, int(maxlen*1.2) + 2)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

def compress_group_label(s: str) -> str:
    parts = [p.strip() for p in str(s).split(";") if p.strip()]
    if len(parts) == 2:
        def split_pref_suf(t):
            t = t.replace(" ", "")
            i = t.rfind("-")
            if i == -1: return None, None
            return t[:i+1], t[i+1:]
        p1, p2 = split_pref_suf(parts[0])
        q1, q2 = split_pref_suf(parts[1])
        if p1 and q1 and (p1 == q1) and re.fullmatch(r"\d+", p2 or "") and re.fullmatch(r"\d+", q2 or ""):
            return f"{p1}{p2}/{q2}"
    return " / ".join(parts) if parts else str(s)

# ===== Основная функция =====
def process_step2_and_plot(xlsx_from_step1: bytes):
    """
    Вход: байты XLSX из Шага 1.
    Выход:
      - xlsx2_bytes: XLSX «…, группы помечены цветом.xlsx»,
      - charts: список [(имя_листа, png_bytes)].
    """
    # ---- Чтение книги ----
    bio = BytesIO(xlsx_from_step1)
    xls = pd.ExcelFile(bio, engine="openpyxl")

    raw_sheets, proc_sheets, all_groups = {}, {}, set()
    COL_DATE_VARS   = ["Дата","дата","Даты","даты"]
    COL_START_VARS  = ["Начало","start","время начала","начало"]
    COL_END_VARS    = ["Конец","end","время конца","конец"]
    COL_GROUPS_VARS = ["Группы","Группа","groups","группы","группа"]
    COL_TYPE_VARS   = ["Тип","тип","type","Тип занятия","вид занятия","форма занятия","формат","занятие","занятия"]

    def is_lecture(val):
        if pd.isna(val): return False
        s = str(val).strip().lower()
        return any([s.startswith("лек"), "лекц" in s, s=="лк", "lecture" in s, s=="lec"])

    for sheet in xls.sheet_names:
        try:
            df_raw = pd.read_excel(xls, sheet_name=sheet, dtype=object)
        except Exception as e:
            continue

        raw_sheets[sheet] = df_raw.copy()
        cols = list(df_raw.columns)
        c_date   = find_column(cols, COL_DATE_VARS)
        c_start  = find_column(cols, COL_START_VARS)
        c_end    = find_column(cols, COL_END_VARS)
        c_groups = find_column(cols, COL_GROUPS_VARS)
        c_type   = find_column(cols, COL_TYPE_VARS)

        missing = [name for name, col in [("Дата", c_date),("Начало", c_start),("Конец", c_end),("Группы", c_groups)] if col is None]
        if missing:
            # пропускаем листы, где не нашли базовые колонки
            continue

        df = df_raw.copy()
        df["_Дата"]    = to_date_series(df[c_date]).dt.date
        df["_Начало"]  = df[c_start].apply(parse_time_cell)
        df["_Конец"]   = df[c_end].apply(parse_time_cell)
        df["_Тип"]     = df[c_type] if c_type is not None else None
        df["_Лекция"]  = df["_Тип"].apply(is_lecture) if c_type is not None else False

        def fix_order(row):
            t1, t2 = row["_Начало"], row["_Конец"]
            if t1 and t2 and (t2.hour*60+t2.minute) < (t1.hour*60+t1.minute):
                row["_Начало"], row["_Конец"] = t2, t1
            return row
        df = df.apply(fix_order, axis=1)

        df["_Группы"] = df[c_groups].astype(str).apply(lambda s: re.sub(r"\s+", " ", s.strip()))
        df["_ГруппыКорот"] = df["_Группы"].apply(compress_group_label)

        iso_wday = df["_Дата"].apply(lambda d: None if pd.isna(d) else datetime(d.year, d.month, d.day).weekday())
        df["_widx"] = iso_wday
        df["_ДеньНедели"] = [RU_WEEKDAYS[w] if (w is not None) else None for w in iso_wday]

        mask_ok = (~df["_Дата"].isna()) & df["_Начало"].notna() & df["_Конец"].notna() & df["_Группы"].astype(bool)
        df = df.loc[mask_ok].drop_duplicates(subset=["_Дата","_Начало","_Конец","_Группы","_Лекция"]).copy()
        df["_m_start"] = df["_Начало"].apply(lambda t: t.hour*60 + t.minute if t else None)
        df["_m_end"]   = df["_Конец"].apply(lambda t: t.hour*60 + t.minute if t else None)

        all_groups.update(df["_Группы"].unique())
        proc_sheets[sheet] = df

    if not proc_sheets:
        raise RuntimeError("Нет листов с валидной структурой. Убедитесь, что вход — результат Шага 1.")

    # ===== Глобальный старт «учебной зебры» для всей книги =====
    global_first_dates = [df["_Дата"].min() for df in proc_sheets.values() if not df.empty]
    global_first_dates = [d for d in global_first_dates if pd.notna(d)]
    global_first_mon = week_monday(min(global_first_dates)) if global_first_dates else None

    # ===== Палитра =====
    def hex_to_rgb01(h):
        h = h.strip().lstrip('#'); return (int(h[0:2],16)/255.0, int(h[2:4],16)/255.0, int(h[4:6],16)/255.0)
    def rgb_to_hex(c): return "#{:02X}{:02X}{:02X}".format(int(c[0]*255), int(c[1]*255), int(c[2]*255))
    def srgb_to_linear(c): return c/12.92 if c <= 0.04045 else ((c+0.055)/1.055)**2.4
    def rgb_to_lab(rgb):
        r,g,b = (srgb_to_linear(x) for x in rgb)
        X = 0.4124564*r + 0.3575761*g + 0.1804375*b
        Y = 0.2126729*r + 0.7151522*g + 0.0721750*b
        Z = 0.0193339*r + 0.1191920*g + 0.9503041*b
        Xn, Yn, Zn = 0.95047, 1.0, 1.08883
        x, y, z = X/Xn, Y/Yn, Z/Zn
        def f(t): return t**(1/3) if t > 0.008856 else (7.787*t + 16/116)
        fx, fy, fz = f(x), f(y), f(z)
        return (116*fy - 16, 500*(fx - fy), 200*(fy - fz))
    def delta_e76(l1, l2): return math.sqrt((l1[0]-l2[0])**2 + (l1[1]-l2[1])**2 + (l1[2]-l2[2])**2)

    CURATED = [
        "F3C300","875692","F38400","A1CAF1","BE0032","C2B280","008856","E68FAC","0067A5","F99379","604E97","F6A600",
        "B3446C","DCD300","882D17","8DB600","E25822","2B3D26",
        "1F77B4","FF7F0E","2CA02C","D62728","9467BD","8C564B","E377C2","BCBD22","17BECF",
        "AEC7E8","FFBB78","98DF8A","FF9896","C5B0D5","C49C94","F7B6D2","DBDB8D","9EDAE5",
        "0072B2","E69F00","56B4E9","D55E00","009E73","F0E442","CC79A7",
        "F44336","E91E63","9C27B0","673AB7","3F51B5","2196F3","03A9F4","00BCD4","009688","4CAF50",
        "8BC34A","CDDC39","FFEB3B","FFC107","FF9800","FF5722","795548","607D8B"
    ]
    CURATED = [hex_to_rgb01(h) for h in CURATED]
    labs = [rgb_to_lab(c) for c in CURATED]
    unique_groups = sorted(list(all_groups), key=lambda s: s.lower())
    palette_list = []
    thr = 22.0
    for c,lab in zip(CURATED, labs):
        if not palette_list:
            palette_list.append(c)
        else:
            if min(delta_e76(rgb_to_lab(pc), lab) for pc in palette_list) >= thr:
                palette_list.append(c)
        if len(palette_list) >= max(1, len(unique_groups)): break
    while len(palette_list) < len(unique_groups):
        i = len(palette_list)
        h = (i * 0.61803398875) % 1.0
        r,g,b = colorsys.hsv_to_rgb(h, 0.80, 0.96)
        lab = rgb_to_lab((r,g,b))
        if min(delta_e76(rgb_to_lab(pc), lab) for pc in palette_list) >= thr:
            palette_list.append((r,g,b))
    palette = {g: "#{:02X}{:02X}{:02X}".format(int(palette_list[i % len(palette_list)][0]*255),
                                               int(palette_list[i % len(palette_list)][1]*255),
                                               int(palette_list[i % len(palette_list)][2]*255))
               for i,g in enumerate(unique_groups)}

    # ===== Построение диаграмм =====
    charts: List[tuple] = []

    for sheet, df in proc_sheets.items():
        used_wdays = sorted([int(w) for w in df["_widx"].dropna().unique().tolist()])
        if not used_wdays:
            continue

        dmin, dmax = df["_Дата"].min(), df["_Дата"].max()
        x_min = date(dmin.year, dmin.month, 1)
        first_next_month = date(dmax.year, dmax.month, 1) + relativedelta(months=1)
        x_max = first_next_month - timedelta(days=1)

        all_days = [d for d in daterange(x_min, x_max)]
        wd_dates = [d for d in all_days if d.weekday() < 5]
        if not wd_dates:
            continue

        date_to_idx = {d: i for i, d in enumerate(wd_dates)}
        N = len(wd_dates)

        y0, y1 = 8*60, 19*60
        nrows = len(used_wdays)
        ratio = 1.4
        fig_w = FIG_WIDTH_INCH
        fig_h = fig_w / ratio
        fig, axes = plt.subplots(nrows, 1, figsize=(fig_w, fig_h), dpi=FIG_DPI, sharex=True, sharey=True)
        if nrows == 1: axes = [axes]

        # Недельная зебра: с первой занятой недели и далее через одну
        first_activity = df["_Дата"].min()
        week_starts = [d for d in wd_dates if d.weekday()==0 and d >= week_monday(first_activity)]
        for k, mon in enumerate(week_starts):
            if k % 2 == 0:
                left = date_to_idx.get(mon, None)
                if left is None: continue
                right = left + 5
                for ax in axes:
                    ax.axvspan(left, right, facecolor="#000000", alpha=0.04, zorder=0)

        # Границы месяцев
        month_bound_positions = []
        m_cursor_b = date(x_min.year, x_min.month, 1) + relativedelta(months=1)
        while m_cursor_b <= x_max:
            idxs = [date_to_idx[d] for d in wd_dates if d >= m_cursor_b]
            if idxs:
                month_bound_positions.append(min(idxs))
            m_cursor_b += relativedelta(months=1)
        for ax in axes:
            for xline in month_bound_positions:
                ax.axvline(x=xline, linestyle=(0,(6,4)), linewidth=2.0, color="0.2", alpha=0.9, zorder=1)

        boundary_positions = list(range(N+1))
        for ax, w in zip(axes, used_wdays):
            ax.set_xlim(0, N)
            ax.set_ylim(y0, y1)
            ax.xaxis.set_major_locator(FixedLocator(boundary_positions))
            ax.xaxis.set_major_formatter(FixedFormatter([""]*(N+1)))
            ax.grid(which="major", axis="x", linestyle="-", linewidth=0.7, alpha=0.55)

            dates_w = sorted(set(d for d in df.loc[df["_widx"]==w, "_Дата"] if d in date_to_idx))
            tick_positions = [date_to_idx[d] + 0.5 for d in dates_w]
            tick_labels    = [str(d.day) for d in dates_w]

            ax2 = ax.secondary_xaxis('bottom')
            ax2.set_xlim(0, N)
            ax2.xaxis.set_major_locator(FixedLocator(tick_positions))
            ax2.xaxis.set_major_formatter(FixedFormatter(tick_labels))
            ax2.tick_params(axis='x', length=0, pad=4)
            for lbl, d in zip(ax2.get_xticklabels(), dates_w):
                lbl.set_rotation(0)
                lbl.set_ha("center")
                if d.day % 2 == 0:
                    x, y = lbl.get_position()
                    lbl.set_position((x, y - 0.12))

        for ax in axes:
            yticks = list(range(int(y0), int(y1)+1, 60))
            ax.set_yticks(yticks)
            ax.set_yticklabels([minutes_to_label_custom(m) for m in yticks])
            for t in ax.get_yticklabels():
                t.set_ha("left")
            ylabels = ax.get_yticklabels()
            base_pad = 18
            if ylabels:
                fs = ylabels[0].get_fontsize()
                char_w = int(round(fs * 0.60))
                ax.tick_params(axis='y', pad=base_pad + char_w)
            else:
                ax.tick_params(axis='y', pad=base_pad)
            ax.grid(which="major", axis="y", linestyle="-", linewidth=0.5, alpha=0.35)
            for spine in ["top","right","left","bottom"]:
                ax.spines[spine].set_visible(False)

        ax_top = axes[0]
        ax_bottom_axes = axes[-1]
        m_cursor = date(x_min.year, x_min.month, 1)
        while m_cursor <= x_max:
            next_month = (m_cursor + relativedelta(months=1))
            month_wd = [d for d in wd_dates if m_cursor <= d < next_month]
            if month_wd:
                mid = (date_to_idx[month_wd[0]] + date_to_idx[month_wd[-1]] + 1)/2.0
                ax_top.annotate(month_name_ru(datetime(m_cursor.year, m_cursor.month, 1)),
                                xy=(mid, 1), xycoords=('data','axes fraction'),
                                xytext=(0, 6), textcoords='offset points',
                                ha="center", va="bottom", fontsize=12, clip_on=False)
                ax_bottom_axes.annotate(month_name_ru(datetime(m_cursor.year, m_cursor.month, 1)),
                                        xy=(mid, 0), xycoords=('data','axes fraction'),
                                        xytext=(0, -18), textcoords='offset points',
                                        ha="center", va="top", fontsize=12, clip_on=False)
            m_cursor = next_month

        for ax, w in zip(axes, used_wdays):
            ax.text(0.005, 0.5, RU_WEEKDAYS[w], transform=ax.transAxes,
                    ha="left", va="center", fontsize=12)

        first_pos = {}
        for ax, w in zip(axes, used_wdays):
            dfx = df[df["_widx"] == w].copy()
            if dfx.empty: continue
            key_counts = {}
            dfx = dfx.sort_values(["_Дата","_m_start","_m_end"])
            for _, row in dfx.iterrows():
                d0 = row["_Дата"]
                if d0 not in date_to_idx:
                    continue
                idx = date_to_idx[d0]
                x_center = idx + 0.5
                start_m = int(row["_m_start"]); end_m = int(row["_m_end"])
                grp = row["_Группы"]; grp_short = row["_ГруппыКорот"]; is_lec = bool(row["_Лекция"])
                kkey = (idx, start_m, end_m); cnt = key_counts.get(kkey, 0); key_counts[kkey] = cnt + 1
                x_off = sym_offset(cnt, 0.03)
                x_left = x_center - (BAR_WIDTH_DAYS/2.0) + x_off

                rect = mpatches.Rectangle((x_left, start_m), BAR_WIDTH_DAYS, max(1, end_m - start_m),
                                          linewidth=(0.9 if is_lec else 0.0),
                                          edgecolor=("black" if is_lec else "none"),
                                          facecolor=palette.get(grp, "#999999"),
                                          alpha=0.95)
                ax.add_patch(rect)

                if grp not in first_pos:
                    y_mid = (start_m + end_m) / 2.0
                    first_pos[grp] = (ax, x_left + BAR_WIDTH_DAYS + 0.05, y_mid, grp_short)

        for grp, (ax, x_text, y_text, label_text) in first_pos.items():
            txt = ax.text(x_text, y_text, label_text, fontsize=10, va="center", ha="left", zorder=10, clip_on=True)
            txt.set_path_effects([path_effects.Stroke(linewidth=2.2, foreground="white"), path_effects.Normal()])

        width_frac  = 0.95
        height_frac = 0.90
        left   = (1 - width_frac)  / 2
        right  = 1 - left
        bottom = (1 - height_frac) / 2
        top    = 1 - bottom
        plt.subplots_adjust(left=left, right=right, top=top, bottom=bottom, hspace=0.10)

        png_buf = BytesIO()
        suffix = "" if len(proc_sheets) == 1 else f" — {sheet}"
        fig.savefig(png_buf, dpi=FIG_DPI, bbox_inches="tight", transparent=False, format="png")
        plt.close(fig)
        charts.append((sheet, png_buf.getvalue()))

    # ===== XLSX с раскраской групп и «зеброй» недель =====
    legend_df = pd.DataFrame({"Группы": sorted(unique_groups, key=lambda s: s.lower()),
                              "Цвет": [palette[g] for g in sorted(unique_groups, key=lambda s: s.lower())]})
    out_buf = BytesIO()
    with pd.ExcelWriter(out_buf, engine="openpyxl") as writer:
        # исходные листы (из входной книги)
        for sheet in xls.sheet_names:
            df_raw = raw_sheets.get(sheet)
            if df_raw is not None:
                df_raw.to_excel(writer, sheet_name=sheet, index=False)
        # лист Legend
        legend_df.to_excel(writer, sheet_name="Legend", index=False)

        wb = writer.book
        light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # ---- раскраска рабочих листов ----
        for sheet in xls.sheet_names:
            ws = writer.sheets.get(sheet)
            df_raw = raw_sheets.get(sheet)
            if ws is None or df_raw is None or ws.max_row < 2:
                continue

            # Нормализация и поиск нужных колонок
            g_col_name = find_column(list(df_raw.columns), ["Группы","Группа","groups","группы","группа"])
            d_col_name = find_column(list(df_raw.columns), ["Дата","дата","Даты","даты"])

            # Индекс колонки групп
            g_col_idx = None
            if g_col_name is not None:
                try:
                    g_col_idx = list(df_raw.columns).index(g_col_name) + 1
                except ValueError:
                    g_col_idx = None

            # Палитра групп — закраска ячеек группы
            if g_col_idx is not None:
                for r in range(2, ws.max_row+1):
                    val = ws.cell(row=r, column=g_col_idx).value
                    key = re.sub(r"\s+", " ", str(val).strip()) if val is not None else ""
                    hexcol = palette.get(key)
                    if hexcol:
                        rgb = hexcol.lstrip("#").upper()
                        ws.cell(row=r, column=g_col_idx).fill = PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")

            # «Зебра» учебных недель от глобальной первой недели
            if (d_col_name is not None) and (global_first_mon is not None):
                ser_dates = to_date_series(df_raw[d_col_name]).dt.date
                for i_excel, d in enumerate(ser_dates, start=2):
                    if pd.isna(d):
                        continue
                    k = (week_monday(d) - global_first_mon).days // 7
                    if k >= 0 and (k % 2 == 0):
                        for c in range(1, ws.max_column+1):
                            if g_col_idx is not None and c == g_col_idx:
                                continue
                            ws.cell(row=i_excel, column=c).fill = light_gray

        # ---- раскраска в Legend ----
        ws_leg = writer.sheets.get("Legend")
        if ws_leg is not None and ws_leg.max_row >= 2:
            for r in range(2, ws_leg.max_row+1):
                key = ws_leg.cell(row=r, column=1).value
                key = re.sub(r"\s+", " ", str(key).strip()) if key is not None else ""
                hexcol = palette.get(key)
                if hexcol:
                    rgb = hexcol.lstrip("#").upper()
                    ws_leg.cell(row=r, column=1).fill = PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")

        # ---- автоширина ----
        all_for_width = {name: raw_sheets[name] for name in xls.sheet_names if name in raw_sheets}
        all_for_width["Legend"] = legend_df
        autosize_all_sheets(writer, all_for_width)

    return out_buf.getvalue(), charts
